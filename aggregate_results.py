#!/usr/bin/env python3
"""
Aggregate metadata survey results from multiple JSON files.
Reads all JSON files from 'Responses (JSON)' folder and outputs an Excel file
with average votes for each preset and metadata tag.
"""

import json
from pathlib import Path
from collections import defaultdict
from statistics import mean

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Configuration
RESPONSES_FOLDER = "Responses (JSON)"
OUTPUT_FILE = "aggregated_results.xlsx"

def load_all_responses(folder_path):
    """Load all JSON response files from the specified folder."""
    responses = []
    folder = Path(folder_path)

    if not folder.exists():
        raise FileNotFoundError(f"Folder '{folder_path}' not found")

    json_files = list(folder.glob("*.json"))

    if not json_files:
        raise FileNotFoundError(f"No JSON files found in '{folder_path}'")

    print(f"Found {len(json_files)} JSON file(s)")

    for json_file in json_files:
        with open(json_file, 'r') as f:
            data = json.load(f)
            responses.append(data)
            print(f"  - Loaded {json_file.name}")

    return responses

def aggregate_votes(all_responses):
    """
    Aggregate votes across all respondents.
    Returns a dictionary: {preset_name: {tag: [list of votes]}}
    """
    aggregated = defaultdict(lambda: defaultdict(list))

    for response_data in all_responses:
        for preset_response in response_data['responses']:
            preset_name = preset_response['presetName']

            for vote_key, vote_value in preset_response['votes'].items():
                aggregated[preset_name][vote_key].append(vote_value)

    return aggregated

def calculate_averages(aggregated_data):
    """
    Calculate average votes for each preset and tag.
    Returns a dictionary: {preset_name: {tag: average_vote}}
    """
    averages = {}

    for preset_name, tags in aggregated_data.items():
        averages[preset_name] = {}
        for tag, votes in tags.items():
            averages[preset_name][tag] = round(mean(votes), 2)

    return averages

def extract_category(tag):
    """Extract category from a tag name (e.g., 'advancedInstrument-Piano' -> 'advancedInstrument')."""
    if '-' in tag:
        return tag.split('-')[0]
    return 'other'

def get_gradient_color(value):
    """
    Get a gradient color based on value (0-5 scale).
    - 0-3: Red gradient (darkest red at 0, lightest red approaching 3)
    - 3-5: Green gradient (lightest green just above 3, darkest green at 5)
    Returns a PatternFill object.
    """
    if not isinstance(value, (int, float)):
        return None

    # Clamp value between 0 and 5
    value = max(0, min(5, value))

    if value < 3:
        # Red gradient: 0 (dark red) to 3 (light red/white)
        # Interpolate from dark red (CC0000) to light pink (FFE6E6)
        ratio = value / 3.0  # 0 at value=0, 1 at value=3

        # Start: Dark red (204, 0, 0)
        # End: Light pink (255, 230, 230)
        r = int(204 + (255 - 204) * ratio)
        g = int(0 + (230 - 0) * ratio)
        b = int(0 + (230 - 0) * ratio)

    else:
        # Green gradient: 3 (light green/white) to 5 (dark green)
        # Interpolate from light green (E6FFE6) to dark green (00CC00)
        ratio = (value - 3.0) / 2.0  # 0 at value=3, 1 at value=5

        # Start: Light green (230, 255, 230)
        # End: Dark green (0, 204, 0)
        r = int(230 + (0 - 230) * ratio)
        g = int(255 + (204 - 255) * ratio)
        b = int(230 + (0 - 230) * ratio)

    # Convert to hex
    hex_color = f'{r:02X}{g:02X}{b:02X}'

    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def calculate_category_averages(preset_data):
    """
    Calculate average of averages for each metadata category.
    Returns a dictionary: {category: average_of_averages}
    """
    category_values = defaultdict(list)

    # Group values by category
    for tag, avg_value in preset_data.items():
        category = extract_category(tag)
        category_values[category].append(avg_value)

    # Calculate average for each category
    category_averages = {}
    for category, values in category_values.items():
        category_averages[category] = round(mean(values), 2)

    return category_averages

def write_excel(averages, output_file):
    """Write aggregated results to Excel file with custom columns per preset and bold headers."""

    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")

    # Create workbook and get active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Aggregated Results"

    # Sort preset names
    preset_names = sorted(averages.keys())

    # Bold font for headers
    bold_font = Font(bold=True)

    # Current row tracker
    current_row = 1

    # Store category averages for summary section
    all_category_averages = {}

    # Process each preset
    for preset_name in preset_names:
        preset_data = averages[preset_name]

        # Get sorted tags for this preset
        tags = sorted(preset_data.keys())

        # Calculate category averages
        category_averages = calculate_category_averages(preset_data)
        all_category_averages[preset_name] = category_averages
        category_names = sorted(category_averages.keys())

        # Write header row for this preset
        header = ['Preset Name'] + tags + [f'{cat}_AVG' for cat in category_names]
        for col_idx, header_value in enumerate(header, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header_value)
            cell.font = bold_font

        current_row += 1

        # Write data row for this preset
        row = [preset_name] + [preset_data[tag] for tag in tags] + [category_averages[cat] for cat in category_names]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)

        current_row += 1

        # Write blank spacer row
        current_row += 1

    # Add extra spacer rows before summary section
    current_row += 2

    # Write summary section header
    ws.cell(row=current_row, column=1, value="CATEGORY AVERAGES SUMMARY").font = bold_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="(For easy cross-preset comparison)").font = bold_font
    current_row += 2

    # Get all unique categories across all presets
    all_categories = set()
    for cat_avgs in all_category_averages.values():
        all_categories.update(cat_avgs.keys())
    all_categories = sorted(all_categories)

    # Write summary table header
    summary_header = ['Preset Name'] + [f'{cat}_AVG' for cat in all_categories]
    for col_idx, header_value in enumerate(summary_header, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header_value)
        cell.font = bold_font

    current_row += 1

    # Write summary table data
    for preset_name in preset_names:
        category_averages = all_category_averages[preset_name]
        row_data = [preset_name]

        # Add category averages (use empty string if category not present for this preset)
        for cat in all_categories:
            row_data.append(category_averages.get(cat, ''))

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)

            # Apply gradient coloring to numeric values (skip preset name column)
            if col_idx > 1 and isinstance(value, (int, float)):
                gradient_fill = get_gradient_color(value)
                if gradient_fill:
                    cell.fill = gradient_fill

        current_row += 1

    # Save workbook
    wb.save(output_file)

    print(f"\n✓ Results written to '{output_file}'")
    print(f"  - {len(preset_names)} presets")
    print(f"  - Each preset with its own custom columns")
    print(f"  - Category averages included for each preset")
    print(f"  - Summary section added at bottom for easy comparison")
    print(f"  - Headers are bold for easy reading")
    print(f"  - Gradient coloring applied: Red (0) → White (3) → Green (5)")

def main():
    """Main execution function."""
    try:
        print("=" * 60)
        print("Metadata Survey Results Aggregator")
        print("=" * 60)
        print()

        # Load all JSON responses
        all_responses = load_all_responses(RESPONSES_FOLDER)

        # Aggregate votes
        print("\nAggregating votes...")
        aggregated = aggregate_votes(all_responses)

        # Calculate averages
        print("Calculating averages...")
        averages = calculate_averages(aggregated)

        # Write to Excel
        print("Writing Excel file...")
        write_excel(averages, OUTPUT_FILE)

        print("\n" + "=" * 60)
        print("✓ Aggregation complete!")
        print("=" * 60)

    except Exception as e:
        print(f"\n✗ Error: {e}")
        return 1

    return 0

if __name__ == "__main__":
    exit(main())
